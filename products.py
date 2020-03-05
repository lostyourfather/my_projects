from openpyxl import load_workbook
wb = load_workbook("products_sheet.xlsx")
sheet1 = wb['Справочник']
sheet2 = wb['Раскладка']
kal = {}
bel = {}
jir = {}
ugl = {}
def firstList(kal,bel,jir,ugl):
    names = []
    for i in range(2, sheet1.max_row+1):
        name = sheet1.cell(row=i, column=1).value
        names.append(name)
        kal[name] = sheet1.cell(row=i, column=2).value
        bel[name] = sheet1.cell(row=i, column=3).value
        jir[name] = sheet1.cell(row=i, column=4).value
        ugl[name] = sheet1.cell(row=i, column=5).value
    return kal,bel,jir,ugl
firstList(kal,bel,jir,ugl)
day1 = [0,0,0,0]
day2 = [0,0,0,0]
day3 = [0,0,0,0]
day4 = [0,0,0,0]
day5 = [0,0,0,0]
day6 = [0,0,0,0]
day7 = [0,0,0,0]
day8 = [0,0,0,0]
day9 = [0,0,0,0]
for j in range(2, sheet2.max_row+1):
    productName = sheet2.cell(row=j, column=2).value
    productVes = sheet2.cell(row=j, column=3).value
    if sheet2.cell(row=j, column=1).value == 1:
        day1[0] = day1[0] + ((kal[productName]/100) * productVes)
        day1[1] = day1[1] + ((bel[productName]/100) * productVes)
        day1[2] = day1[2] + ((jir[productName]/100) * productVes)
        day1[3] = day1[3] + ((ugl[productName]/100) * productVes)
    if sheet2.cell(row=j, column=1).value == 2:
        day2[0] = day2[0] + ((kal[productName]/100) * productVes)
        day2[1] = day2[1] + ((bel[productName] / 100) * productVes)
        day2[2] = day2[2] + ((jir[productName] / 100) * productVes)
        day2[3] = day2[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 3:
        day3[0] = day3[0] + ((kal[productName]/100) * productVes)
        day3[1] = day3[1] + ((bel[productName] / 100) * productVes)
        day3[2] = day3[2] + ((jir[productName] / 100) * productVes)
        day3[3] = day3[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 4:
        day4[0] = day4[0] + ((kal[productName]/100) * productVes)
        day4[1] = day4[1] + ((bel[productName] / 100) * productVes)
        day4[2] = day4[2] + ((jir[productName] / 100) * productVes)
        day4[3] = day4[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 5:
        day5[0] = day5[0] + ((kal[productName]/100) * productVes)
        day5[1] = day5[1] + ((bel[productName] / 100) * productVes)
        day5[2] = day5[2] + ((jir[productName] / 100) * productVes)
        day5[3] = day5[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 6:
        day6[0] = day6[0] + ((kal[productName]/100) * productVes)
        day6[1] = day6[1] + ((bel[productName] / 100) * productVes)
        day6[2] = day6[2] + ((jir[productName] / 100) * productVes)
        day6[3] = day6[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 7:
        day7[0] = day7[0] + ((kal[productName]/100) * productVes)
        day7[1] = day7[1] + ((bel[productName] / 100) * productVes)
        day7[2] = day7[2] + ((jir[productName] / 100) * productVes)
        day7[3] = day7[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 8:
        day8[0] = day8[0] + ((kal[productName]/100) * productVes)
        day8[1] = day8[1] + ((bel[productName] / 100) * productVes)
        day8[2] = day8[2] + ((jir[productName] / 100) * productVes)
        day8[3] = day8[3] + ((ugl[productName] / 100) * productVes)
    if sheet2.cell(row=j, column=1).value == 9:
        day9[0] = day9[0] + ((kal[productName]/100) * productVes)
        day9[1] = day9[1] + ((bel[productName] / 100) * productVes)
        day9[2] = day9[2] + ((jir[productName] / 100) * productVes)
        day9[3] = day9[3] + ((ugl[productName] / 100) * productVes)

print(*day1)
print(*day2)
print(*day3)
print(*day4)
print(*day5)
print(*day6)
print(*day7)
print(*day8)
print(*day9)
